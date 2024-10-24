import pandas as pd
import json
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load the raw data from 'data/raw.xlsx'
raw_data_path = 'data/raw.xlsx'
sheet_s8 = 'S8 | Modeling vs "raw" database'
sheet_s2 = 'S2 | Ions'

# Load the first four columns from 'S8 | Modeling vs "raw" database' sheet: Dataset ID, IL ID, Cation, Anion, T / K, and η / mPa s
df_s8 = pd.read_excel(raw_data_path, sheet_name=sheet_s8, usecols=[0, 1, 2, 3, 8, 9])

# Load the column headers for functional groups from 'S2 | Ions' starting from the 9th column
df_s2 = pd.read_excel(raw_data_path, sheet_name=sheet_s2)
functional_group_headers = df_s2.columns[8:].tolist()

# Prepare the data structure
new_df_columns = ['Dataset ID', 'IL ID', 'Cation', 'Anion', 'T / K', 'η / mPa s', 'a', 'b'] + functional_group_headers

# Load the functional group data from 'ionic_liquid_info.json'
json_data_path = 'generate-fg-graph/ionic_liquid_info.json'
with open(json_data_path, 'r') as f:
    ionic_liquid_data = json.load(f)

# Create a new workbook and add a sheet
wb = Workbook()
ws = wb.active
ws.title = 'Dataset'

# Write the header row
ws.append(new_df_columns)

# Function to write rows in chunks
def write_chunked_data(ws, data_chunk):
    for row in data_chunk:
        ws.append(row)

# Prepare the data for chunk writing
data_chunk = []
for il_info in tqdm(ionic_liquid_data, desc="Processing Ionic Liquids", unit="liquid"):
    il_id = il_info['IL ID']
    cation_fg = il_info.get('Cation Functional Groups', {})
    anion_fg = il_info.get('Anion Functional Groups', {})

    combined_fg = {}
    for fg, count in cation_fg.items():
        combined_fg[fg] = combined_fg.get(fg, 0) + count
    for fg, count in anion_fg.items():
        combined_fg[fg] = combined_fg.get(fg, 0) + count

    # Find all rows in df_s8 that match the current IL ID
    matching_rows = df_s8[df_s8['IL ID'] == il_id]

    # Iterate over all matching rows
    for _, df_row in matching_rows.iterrows():
        # Prepare row data for each match
        row_data = [df_row['Dataset ID'], df_row['IL ID'], df_row['Cation'], df_row['Anion'],
                    df_row['T / K'], df_row['η / mPa s'], None, None]  # Placeholder for 'a' and 'b'

        # Add functional group counts
        for fg in functional_group_headers:
            row_data.append(combined_fg.get(fg, 0))

        data_chunk.append(row_data)

    # Write in chunks of 100 rows at a time to avoid memory issues
    if len(data_chunk) >= 100:
        write_chunked_data(ws, data_chunk)
        data_chunk = []  # Reset chunk

# Write any remaining data
if data_chunk:
    write_chunked_data(ws, data_chunk)

# Save the file
output_path = 'data/working-dataset.xlsx'
wb.save(output_path)

print(f"Data saved to {output_path}")

# Load the workbook and worksheet to apply formatting
wb = load_workbook(output_path)
ws = wb['Dataset']

# Define formatting styles: bold red font and yellow fill
red_bold_font = Font(color="FF0000", bold=True)
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Create a progress bar for the outer loop (iterating over rows)
for row in tqdm(ws.iter_rows(min_row=2, min_col=9, max_col=ws.max_column, max_row=ws.max_row),desc="Formatting Cells", unit="row", total=ws.max_row - 1):
    for cell in row:
        if isinstance(cell.value, (int, float)) and cell.value != 0:
            cell.font = red_bold_font
            cell.fill = yellow_fill

# Save the formatted workbook
wb.save(output_path)

print(f"Formatted Excel file saved at: {output_path}")
